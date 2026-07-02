"""
excel_exporter.py
Generates Excel workbook matching PKm_Rule_Curve-model.xlsx structure exactly.

7 sheets:
  1_Summary            — reservoir params + 10-case cushion table
  2_Demand             — 36-period demand with ALR/TMT split (if available)
  3_Inflows            — P50 & P90 inflows with Inf/Demand ratio & status
  4_UpperRC_AllCases   — URC levels for all 10 cases (36 rows × 12 cols)
  5_LowerRC_AllCases   — LRC levels for all 10 cases (36 rows × 12 cols)
  6_DemSat_AllCases    — demand satisfaction % for all 10 cases
  10_CrossComparison   — URC levels cross-comparison table
"""

import numpy as np
import pandas as pd
from io import BytesIO
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ── colour palette (hex, no #) ────────────────────────────────────────
C_HEADER_BLUE   = "0055A5"
C_HEADER_DARK   = "1F3864"
C_HEADER_ORANGE = "FF6600"
C_SUBHDR_BLUE   = "BDD7EE"
C_SUBHDR_GREY   = "D9D9D9"
C_SUBHDR_GREEN  = "C6EFCE"
C_ROW_ALT       = "F2F7FB"
C_ROW_WHITE     = "FFFFFF"
C_CUSHION_ZONE  = "DDEEFF"   # Aug-III to Oct-III
C_ANCHOR        = "FFE0E0"   # Nov-II / May-III
C_FRL           = "FFF2CC"   # FRL reference
C_GREEN_FILL    = "C6EFCE"
C_YELLOW_FILL   = "FFEB9C"
C_RED_FILL      = "FFC7CE"
C_WHITE         = "FFFFFF"

THIN = Side(style='thin', color='AAAAAA')
MED  = Side(style='medium', color='555555')
BORDER_THIN  = Border(left=THIN,  right=THIN,  top=THIN,  bottom=THIN)
BORDER_MED   = Border(left=MED,   right=MED,   top=MED,   bottom=MED)


def _hfill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color="000000", size=9, name="Calibri"):
    return Font(bold=bold, color=color, size=size, name=name)

def _align(h='center', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _write(ws, row, col, value, bold=False, bg=None, fg="000000",
           halign='center', wrap=False, size=9, border=True, num_fmt=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = _font(bold=bold, color=fg, size=size)
    cell.alignment = _align(h=halign, wrap=wrap)
    if bg:
        cell.fill  = _hfill(bg)
    if border:
        cell.border = BORDER_THIN
    if num_fmt:
        cell.number_format = num_fmt
    return cell

def _header_row(ws, row, headers, bgs, widths=None):
    """Write a styled header row."""
    for c, (hdr, bg) in enumerate(zip(headers, bgs), 1):
        _write(ws, row, c, hdr, bold=True, bg=bg, fg="FFFFFF",
               halign='center', wrap=True)
    if widths:
        for c, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(c)].width = w

def _set_col_widths(ws, widths):
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w

def _freeze(ws, cell="A3"):
    ws.freeze_panes = cell

PERIODS = [
    'Jul-I','Jul-II','Jul-III',
    'Aug-I','Aug-II','Aug-III',
    'Sep-I','Sep-II','Sep-III',
    'Oct-I','Oct-II','Oct-III',
    'Nov-I','Nov-II','Nov-III',
    'Dec-I','Dec-II','Dec-III',
    'Jan-I','Jan-II','Jan-III',
    'Feb-I','Feb-II','Feb-III',
    'Mar-I','Mar-II','Mar-III',
    'Apr-I','Apr-II','Apr-III',
    'May-I','May-II','May-III',
    'Jun-I','Jun-II','Jun-III',
]
CUSHION_PIDS = range(5, 12)   # Aug-III to Oct-III
UPPER_ANCHOR = 13              # Nov-II
LOWER_ANCHOR = 32              # May-III


# ══════════════════════════════════════════════════════════════════════
#  SHEET 1: Summary
# ══════════════════════════════════════════════════════════════════════
def _sheet_summary(wb, reservoir_name, river_name, frl_m, mddl_m,
                   frl_s, mddl_s, all_cases, unit_label):
    ws = wb.create_sheet("1_Summary")
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:L1")
    c = ws.cell(row=1, column=1,
                value=f"{reservoir_name.upper()} — RULE CURVE DESIGN  |  Units: {unit_label}")
    c.font = _font(bold=True, color="FFFFFF", size=12)
    c.fill = _hfill(C_HEADER_BLUE)
    c.alignment = _align(h='center')

    ws.merge_cells("A2:L2")
    c2 = ws.cell(row=2, column=1,
                 value=f"River: {river_name}  |  Backward calculation method (CWC)  |  "
                       f"Generated: {datetime.now().strftime('%d-%b-%Y')}")
    c2.font = _font(bold=False, color="FFFFFF", size=9)
    c2.fill = _hfill(C_HEADER_DARK)
    c2.alignment = _align(h='center')

    # Parameter block
    params = [
        ("Full Reservoir Level (FRL)",    f"{frl_m:.2f} m"),
        ("Minimum Drawdown Level (MDDL)", f"{mddl_m:.2f} m"),
        ("FRL Storage",                   f"{frl_s:.0f}"),
        ("MDDL Storage",                  f"{mddl_s:.0f}"),
        ("Live Storage",                  f"{frl_s - mddl_s:.0f}"),
        ("Water Year",                    "Jul-I to Jun-III (36 decades)"),
        ("Upper RC Anchor",               "Nov-II (PID 13)"),
        ("Lower RC Anchor",               "May-III (PID 32)"),
        ("Flood Cushion Zone",            "Aug-III to Oct-III (PID 5–11)"),
        ("No. of Cases",                  str(len(all_cases))),
    ]
    ws.merge_cells("A4:B4")
    _write(ws, 4, 1, "RESERVOIR PARAMETERS", bold=True, bg=C_SUBHDR_BLUE, fg="000000",
           halign='center', size=10)
    ws.cell(row=4, column=2)  # merged

    for i, (param, val) in enumerate(params, 5):
        _write(ws, i, 1, param, bold=True, bg=C_ROW_ALT, halign='left', size=9)
        _write(ws, i, 2, val,   bold=False, bg=C_ROW_WHITE, halign='left', size=9)

    # Cases table
    row = 17
    ws.merge_cells(f"A{row}:L{row}")
    c = ws.cell(row=row, column=1, value="FLOOD CUSHION CASES")
    c.font = _font(bold=True, color="FFFFFF", size=10)
    c.fill = _hfill(C_HEADER_ORANGE)
    c.alignment = _align(h='center')

    hdrs = ["Case", "Cushion (ft)", "Target Level (m)", "Target Storage",
            "Flood Space", "Flood Risk Score", "Drought Risk", "Recommended"]
    bgs  = [C_HEADER_BLUE]*8
    for ci, h in enumerate(hdrs, 1):
        _write(ws, row+1, ci, h, bold=True, bg=C_HEADER_BLUE, fg="FFFFFF",
               halign='center', wrap=True)

    drought_labels = ['Minimum','Very Low','Low','Low-Moderate','Moderate',
                      'Moderate-High','High','High','Very High','Maximum']
    max_fs = max(c.flood_space_mcm for c in all_cases) or 1

    for i, case in enumerate(all_cases):
        r = row + 2 + i
        bg = C_ROW_ALT if i % 2 == 0 else C_ROW_WHITE
        score = round(case.flood_space_mcm / max_fs * 10, 1)
        rec   = "★ Recommended" if case.case_number == 5 else ""
        vals  = [f"Case {case.case_number}", f"{case.cushion_ft:.1f}",
                 f"{case.target_level_m:.3f}", f"{case.target_storage:.0f}",
                 f"{case.flood_space_mcm:.0f}", f"{score:.1f}",
                 drought_labels[i], rec]
        for ci, v in enumerate(vals, 1):
            bold = (case.case_number == 5)
            _write(ws, r, ci, v, bold=bold,
                   bg="FFF0D0" if bold else bg, halign='center', size=9)

    _set_col_widths(ws, [10,12,16,16,14,14,16,14])
    ws.row_dimensions[1].height = 22
    _freeze(ws, "A3")


# ══════════════════════════════════════════════════════════════════════
#  SHEET 2: Demand
# ══════════════════════════════════════════════════════════════════════
def _sheet_demand(wb, demand, reservoir_name, unit_label):
    ws = wb.create_sheet("2_Demand")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:H1")
    c = ws.cell(row=1, column=1,
                value=f"10-Day Demand — {reservoir_name}  |  Units: {unit_label}/decade")
    c.font = _font(bold=True, color="FFFFFF", size=11)
    c.fill = _hfill(C_HEADER_BLUE)
    c.alignment = _align(h='center')

    hdrs = ["PID","Period","Demand\n(per decade)","Annual\nFraction (%)","Remark"]
    for ci, h in enumerate(hdrs, 1):
        _write(ws, 2, ci, h, bold=True, bg=C_HEADER_BLUE, fg="FFFFFF",
               halign='center', wrap=True)

    annual = sum(demand)
    for pid in range(36):
        r   = pid + 3
        D   = demand[pid]
        bg  = C_CUSHION_ZONE if pid in CUSHION_PIDS else (
              C_ANCHOR if pid in (UPPER_ANCHOR, LOWER_ANCHOR) else
              (C_ROW_ALT if pid % 2 == 0 else C_ROW_WHITE))
        pct = round(D / annual * 100, 1) if annual > 0 else 0
        remark = ("Flood cushion zone" if pid in CUSHION_PIDS else
                  "Upper RC anchor"   if pid == UPPER_ANCHOR else
                  "Lower RC anchor"   if pid == LOWER_ANCHOR else "")
        for ci, v in enumerate([pid, PERIODS[pid], round(D,1), pct, remark], 1):
            _write(ws, r, ci, v, bg=bg, halign='center', size=9)

    # Totals
    rt = 39
    _write(ws, rt, 1, "TOTAL", bold=True, bg=C_SUBHDR_GREY, halign='center')
    _write(ws, rt, 2, "",      bold=True, bg=C_SUBHDR_GREY)
    _write(ws, rt, 3, round(annual,1), bold=True, bg=C_SUBHDR_GREY, num_fmt='#,##0.0')
    _write(ws, rt, 4, 100.0,  bold=True, bg=C_SUBHDR_GREY)
    _write(ws, rt, 5, "",     bold=True, bg=C_SUBHDR_GREY)

    _set_col_widths(ws, [6, 10, 16, 14, 20])
    _freeze(ws, "A3")


# ══════════════════════════════════════════════════════════════════════
#  SHEET 3: Inflows
# ══════════════════════════════════════════════════════════════════════
def _sheet_inflows(wb, q50, q10, demand, reservoir_name, unit_label):
    ws = wb.create_sheet("3_Inflows")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:J1")
    c = ws.cell(row=1, column=1,
                value=f"Inflows — {reservoir_name}  |  50% & 90% Dependable  |  Units: {unit_label}/decade")
    c.font = _font(bold=True, color="FFFFFF", size=11)
    c.fill = _hfill(C_HEADER_BLUE)
    c.alignment = _align(h='center')

    hdrs = ["PID","Period","50% Dep\nInflow","90% Dep\nInflow",
            "Demand","Inf50/\nDemand","Inf90/\nDemand","Status P50","Status P90"]
    for ci, h in enumerate(hdrs, 1):
        _write(ws, 2, ci, h, bold=True, bg=C_HEADER_BLUE, fg="FFFFFF",
               halign='center', wrap=True)

    for pid in range(36):
        r   = pid + 3
        Q50 = q50[pid];  Q90 = q10[pid];  D = demand[pid]
        ratio50 = round(Q50 / D, 2) if D > 0 else 0
        ratio90 = round(Q90 / D, 2) if D > 0 else 0

        def status(r):
            if r >= 1.5: return "Surplus"
            if r >= 0.8: return "Moderate"
            return "Deficit"

        st50 = status(ratio50); st90 = status(ratio90)
        bg = (C_CUSHION_ZONE if pid in CUSHION_PIDS else
              C_ANCHOR if pid in (UPPER_ANCHOR, LOWER_ANCHOR) else
              (C_ROW_ALT if pid % 2 == 0 else C_ROW_WHITE))

        vals = [pid, PERIODS[pid], round(Q50,1), round(Q90,1),
                round(D,1), ratio50, ratio90, st50, st90]
        for ci, v in enumerate(vals, 1):
            s_bg = bg
            if ci == 8:
                s_bg = (C_GREEN_FILL if st50=="Surplus" else
                        C_YELLOW_FILL if st50=="Moderate" else C_RED_FILL)
            elif ci == 9:
                s_bg = (C_GREEN_FILL if st90=="Surplus" else
                        C_YELLOW_FILL if st90=="Moderate" else C_RED_FILL)
            _write(ws, r, ci, v, bg=s_bg, halign='center', size=9)

    # Totals
    rt = 39
    _write(ws, rt, 1, "ANNUAL", bold=True, bg=C_SUBHDR_GREY)
    _write(ws, rt, 2, "",       bold=True, bg=C_SUBHDR_GREY)
    _write(ws, rt, 3, round(sum(q50),0), bold=True, bg=C_SUBHDR_GREY)
    _write(ws, rt, 4, round(sum(q10),0), bold=True, bg=C_SUBHDR_GREY)
    _write(ws, rt, 5, round(sum(demand),0), bold=True, bg=C_SUBHDR_GREY)
    _write(ws, rt, 6, round(sum(q50)/sum(demand),2) if sum(demand) else 0,
           bold=True, bg=C_SUBHDR_GREY)
    for ci in range(7, 10):
        _write(ws, rt, ci, "", bold=True, bg=C_SUBHDR_GREY)

    _set_col_widths(ws, [6,10,14,14,12,12,12,12,12])
    _freeze(ws, "A3")


# ══════════════════════════════════════════════════════════════════════
#  SHEET 4: Upper RC — All Cases
# ══════════════════════════════════════════════════════════════════════
def _sheet_upper_rc(wb, all_cases, q50, demand, unit_label):
    ws = wb.create_sheet("4_UpperRC_AllCases")
    ws.sheet_view.showGridLines = False
    n = len(all_cases)

    ws.merge_cells(f"A1:{get_column_letter(4+n)}1")
    c = ws.cell(row=1, column=1,
                value=f"Upper Rule Curve — All {n} Cases  |  50% Dependable Inflow  |  Units: {unit_label}")
    c.font = _font(bold=True, color="FFFFFF", size=11)
    c.fill = _hfill(C_HEADER_BLUE)
    c.alignment = _align(h='center')

    # Header row
    base_hdrs = ["PID","Period",f"50%Inf\n({unit_label})",f"Demand\n({unit_label})"]
    case_hdrs = [f"C{c.case_number}\n{c.cushion_ft:.1f}ft" for c in all_cases]
    all_hdrs  = base_hdrs + case_hdrs
    for ci, h in enumerate(all_hdrs, 1):
        bg = C_HEADER_BLUE if ci <= 4 else C_HEADER_ORANGE
        _write(ws, 2, ci, h, bold=True, bg=bg, fg="FFFFFF",
               halign='center', wrap=True)

    for pid in range(36):
        r   = pid + 3
        bg_row = (C_CUSHION_ZONE if pid in CUSHION_PIDS else
                  C_ANCHOR if pid == UPPER_ANCHOR else
                  (C_ROW_ALT if pid % 2 == 0 else C_ROW_WHITE))

        base_vals = [pid, PERIODS[pid], round(q50[pid],1), round(demand[pid],1)]
        for ci, v in enumerate(base_vals, 1):
            _write(ws, r, ci, v, bg=bg_row, halign='center', size=9)

        for ci, case in enumerate(all_cases, 5):
            lv = case.upper_rc.loc[case.upper_rc['PID']==pid, 'Level_m'].values[0]
            bg = bg_row
            _write(ws, r, ci, round(lv, 3), bg=bg, halign='center', size=9,
                   num_fmt='0.000')

    # Annual totals row
    rt = 39
    _write(ws, rt, 1, "ANNUAL", bold=True, bg=C_SUBHDR_GREY)
    _write(ws, rt, 2, "", bold=True, bg=C_SUBHDR_GREY)
    _write(ws, rt, 3, round(sum(q50),0), bold=True, bg=C_SUBHDR_GREY)
    _write(ws, rt, 4, round(sum(demand),0), bold=True, bg=C_SUBHDR_GREY)
    for ci in range(5, 5+n):
        _write(ws, rt, ci, "—", bold=True, bg=C_SUBHDR_GREY)

    widths = [6,10,12,12] + [9]*n
    _set_col_widths(ws, widths)
    _freeze(ws, "C3")


# ══════════════════════════════════════════════════════════════════════
#  SHEET 5: Lower RC — All Cases
# ══════════════════════════════════════════════════════════════════════
def _sheet_lower_rc(wb, all_cases, q10, demand, unit_label):
    ws = wb.create_sheet("5_LowerRC_AllCases")
    ws.sheet_view.showGridLines = False
    n = len(all_cases)

    ws.merge_cells(f"A1:{get_column_letter(4+n)}1")
    c = ws.cell(row=1, column=1,
                value=f"Lower Rule Curve — All {n} Cases  |  90% Dependable Inflow  |  Units: {unit_label}")
    c.font = _font(bold=True, color="FFFFFF", size=11)
    c.fill = _hfill(C_HEADER_BLUE)
    c.alignment = _align(h='center')

    base_hdrs = ["PID","Period",f"90%Inf\n({unit_label})",f"Demand\n({unit_label})"]
    case_hdrs = [f"C{c.case_number}\n{c.cushion_ft:.1f}ft" for c in all_cases]
    all_hdrs  = base_hdrs + case_hdrs
    for ci, h in enumerate(all_hdrs, 1):
        bg = C_HEADER_BLUE if ci <= 4 else C_HEADER_DARK
        _write(ws, 2, ci, h, bold=True, bg=bg, fg="FFFFFF",
               halign='center', wrap=True)

    for pid in range(36):
        r   = pid + 3
        bg_row = (C_CUSHION_ZONE if pid in CUSHION_PIDS else
                  C_ANCHOR if pid == LOWER_ANCHOR else
                  (C_ROW_ALT if pid % 2 == 0 else C_ROW_WHITE))

        base_vals = [pid, PERIODS[pid], round(q10[pid],1), round(demand[pid],1)]
        for ci, v in enumerate(base_vals, 1):
            _write(ws, r, ci, v, bg=bg_row, halign='center', size=9)

        # Lower RC is same for all cases (independent of cushion)
        lrc = all_cases[0].lower_rc
        for ci, case in enumerate(all_cases, 5):
            lv = lrc.loc[lrc['PID']==pid, 'Level_m'].values[0]
            _write(ws, r, ci, round(lv, 3), bg=bg_row, halign='center',
                   size=9, num_fmt='0.000')

    rt = 39
    _write(ws, rt, 1, "ANNUAL", bold=True, bg=C_SUBHDR_GREY)
    _write(ws, rt, 2, "", bold=True, bg=C_SUBHDR_GREY)
    _write(ws, rt, 3, round(sum(q10),0), bold=True, bg=C_SUBHDR_GREY)
    _write(ws, rt, 4, round(sum(demand),0), bold=True, bg=C_SUBHDR_GREY)
    for ci in range(5, 5+n):
        _write(ws, rt, ci, "—", bold=True, bg=C_SUBHDR_GREY)

    widths = [6,10,12,12] + [9]*n
    _set_col_widths(ws, widths)
    _freeze(ws, "C3")


# ══════════════════════════════════════════════════════════════════════
#  SHEET 6: Demand Satisfaction — All Cases
# ══════════════════════════════════════════════════════════════════════
def _sheet_demsat(wb, all_cases, q50, demand, frl_s, mddl_s, es_curve, unit_label):
    """
    For each case: Supply = MIN(Q50+available_storage_drawdown, Demand)
    Simplified: Supply = MIN(Q50, Demand) for inflow-based; 
    Full: run forward sim for all 10 cases.
    """
    ws = wb.create_sheet("6_DemSat_AllCases")
    ws.sheet_view.showGridLines = False
    n = len(all_cases)

    ws.merge_cells(f"A1:{get_column_letter(3 + n*2)}1")
    c = ws.cell(row=1, column=1,
                value=f"Demand Satisfaction — All {n} Cases  |  50% Dependable Inflow")
    c.font = _font(bold=True, color="FFFFFF", size=11)
    c.fill = _hfill(C_HEADER_BLUE)
    c.alignment = _align(h='center')

    # Sub-header: base cols
    base_hdrs = ["PID", "Period", f"Demand\n({unit_label})"]
    for ci, h in enumerate(base_hdrs, 1):
        _write(ws, 2, ci, h, bold=True, bg=C_HEADER_BLUE, fg="FFFFFF",
               halign='center', wrap=True)

    # Case headers: Supply + %Sat for each case
    ci = 4
    for case in all_cases:
        ws.merge_cells(f"{get_column_letter(ci)}2:{get_column_letter(ci+1)}2")
        c2 = ws.cell(row=2, column=ci,
                     value=f"Case {case.case_number} ({case.cushion_ft:.1f}ft)")
        c2.font = _font(bold=True, color="FFFFFF", size=9)
        c2.fill = _hfill(C_HEADER_ORANGE)
        c2.alignment = _align(h='center', wrap=True)
        c2.border = BORDER_THIN
        # sub-cols
        _write(ws, 3, ci,   f"Supply\n({unit_label})", bold=True,
               bg=C_SUBHDR_BLUE, halign='center', wrap=True)
        _write(ws, 3, ci+1, "Sat\n(%)",  bold=True,
               bg=C_SUBHDR_GREEN, halign='center', wrap=True)
        ci += 2

    # Forward sim for ALL cases (supply = min(demand, Q50+drawdown from URC))
    # Simple inflow-based: supply = min(Q50, demand) — shows what inflow alone gives
    # We use proper drawdown: supply = URC_storage[t] - URC_storage[t-1] + Q50 capped at demand

    for pid in range(36):
        r     = pid + 4
        D     = demand[pid]
        Q50   = q50[pid]
        bg_row = (C_CUSHION_ZONE if pid in CUSHION_PIDS else
                  C_ANCHOR if pid in (UPPER_ANCHOR, LOWER_ANCHOR) else
                  (C_ROW_ALT if pid % 2 == 0 else C_ROW_WHITE))

        _write(ws, r, 1, pid,         bg=bg_row, halign='center', size=9)
        _write(ws, r, 2, PERIODS[pid],bg=bg_row, halign='center', size=9)
        _write(ws, r, 3, round(D,1),  bg=bg_row, halign='center', size=9)

        ci = 4
        for case in all_cases:
            urc_s = case.upper_rc['Storage_MCM'].values

            # Available from storage drawdown + inflow
            if pid == 0:
                prev_s = frl_s
            else:
                prev_s = urc_s[pid - 1]

            drawdown = max(0, prev_s - urc_s[pid])
            supply   = min(D, Q50 + drawdown)
            sat_pct  = min(100.0, round(supply / D * 100, 1) if D > 0 else 100.0)
            supply   = round(supply, 1)

            # Colour by satisfaction
            sat_bg = (C_GREEN_FILL  if sat_pct >= 90 else
                      C_YELLOW_FILL if sat_pct >= 60 else C_RED_FILL)

            _write(ws, r, ci,   supply,  bg=bg_row, halign='center', size=9)
            _write(ws, r, ci+1, sat_pct, bg=sat_bg, halign='center',
                   size=9, bold=(sat_pct < 60))
            ci += 2

    # Annual totals
    rt = 40
    _write(ws, rt, 1, "ANNUAL", bold=True, bg=C_SUBHDR_GREY)
    _write(ws, rt, 2, "",       bold=True, bg=C_SUBHDR_GREY)
    _write(ws, rt, 3, round(sum(demand),0), bold=True, bg=C_SUBHDR_GREY)
    ci = 4
    for case in all_cases:
        urc_s = case.upper_rc['Storage_MCM'].values
        total_supply = 0
        prev_s = frl_s
        for pid in range(36):
            D    = demand[pid]
            Q50  = q50[pid]
            draw = max(0, prev_s - urc_s[pid])
            total_supply += min(D, Q50 + draw)
            prev_s = urc_s[pid]
        avg_sat = round(total_supply / sum(demand) * 100, 1) if sum(demand) else 0
        _write(ws, rt, ci,   round(total_supply,0), bold=True, bg=C_SUBHDR_GREY)
        _write(ws, rt, ci+1, avg_sat,               bold=True, bg=C_SUBHDR_GREY)
        ci += 2

    widths = [6, 10, 12] + [11, 8]*n
    _set_col_widths(ws, widths)
    ws.row_dimensions[2].height = 28
    ws.row_dimensions[3].height = 28
    _freeze(ws, "D4")


# ══════════════════════════════════════════════════════════════════════
#  SHEET 7: Cross Comparison
# ══════════════════════════════════════════════════════════════════════
def _sheet_cross_comparison(wb, all_cases, demand, unit_label):
    ws = wb.create_sheet("10_CrossComparison")
    ws.sheet_view.showGridLines = False
    n = len(all_cases)

    ws.merge_cells(f"A1:{get_column_letter(4+n)}1")
    c = ws.cell(row=1, column=1,
                value=f"Cross Comparison — Upper Rule Curve Levels  |  All {n} Cases  |  {unit_label}")
    c.font = _font(bold=True, color="FFFFFF", size=11)
    c.fill = _hfill(C_HEADER_BLUE)
    c.alignment = _align(h='center')

    ws.merge_cells(f"A2:{get_column_letter(4+n)}2")
    note = ws.cell(row=2, column=1,
                   value="Flood Cushion Zone (Aug-III–Oct-III) shaded blue  |  "
                         "Nov-II anchor shaded red  |  All values in metres (m)")
    note.font = _font(size=8, color="555555")
    note.fill = _hfill("F8F8F8")
    note.alignment = _align(h='center')

    hdrs = ["PID","Period","FRL\n(m)",f"Demand\n({unit_label})"]
    hdrs += [f"C{c.case_number}\n{c.cushion_ft:.1f}ft" for c in all_cases]
    for ci, h in enumerate(hdrs, 1):
        bg = C_HEADER_BLUE if ci <= 4 else C_HEADER_ORANGE
        _write(ws, 3, ci, h, bold=True, bg=bg, fg="FFFFFF",
               halign='center', wrap=True)

    frl_level = all_cases[0].upper_rc.loc[
        all_cases[0].upper_rc['Storage_MCM'] ==
        all_cases[0].upper_rc['Storage_MCM'].max(), 'Level_m'
    ].values
    frl_m_ref = frl_level[0] if len(frl_level) else 0

    for pid in range(36):
        r = pid + 4
        bg_row = (C_CUSHION_ZONE if pid in CUSHION_PIDS else
                  C_ANCHOR if pid == UPPER_ANCHOR else
                  (C_ROW_ALT if pid % 2 == 0 else C_ROW_WHITE))

        _write(ws, r, 1, pid,             bg=bg_row, halign='center', size=9)
        _write(ws, r, 2, PERIODS[pid],    bg=bg_row, halign='center', size=9)
        _write(ws, r, 3, round(frl_m_ref,3), bg=C_FRL, halign='center',
               size=9, num_fmt='0.000')
        _write(ws, r, 4, round(demand[pid],1), bg=bg_row, halign='center', size=9)

        for ci, case in enumerate(all_cases, 5):
            lv = case.upper_rc.loc[
                case.upper_rc['PID'] == pid, 'Level_m'
            ].values[0]
            _write(ws, r, ci, round(lv,3), bg=bg_row,
                   halign='center', size=9, num_fmt='0.000')

    widths = [6, 10, 9, 12] + [9]*n
    _set_col_widths(ws, widths)
    _freeze(ws, "C4")


# ══════════════════════════════════════════════════════════════════════
#  MAIN BUILDER
# ══════════════════════════════════════════════════════════════════════
def build_excel_workbook(
    reservoir_name: str,
    river_name:     str,
    frl_m:          float,
    mddl_m:         float,
    frl_s:          float,
    mddl_s:         float,
    all_cases,                  # List[RuleCurveCase]
    q50:            np.ndarray,
    q10:            np.ndarray,
    demand:         np.ndarray,
    es_curve,                   # ElevationStorageCurve
    unit_label:     str = "MCM",
) -> bytes:
    """Build complete workbook and return as bytes."""

    wb = Workbook()
    # Remove default empty sheet
    wb.remove(wb.active)

    _sheet_summary(wb, reservoir_name, river_name, frl_m, mddl_m,
                   frl_s, mddl_s, all_cases, unit_label)

    _sheet_demand(wb, demand, reservoir_name, unit_label)

    _sheet_inflows(wb, q50, q10, demand, reservoir_name, unit_label)

    _sheet_upper_rc(wb, all_cases, q50, demand, unit_label)

    _sheet_lower_rc(wb, all_cases, q10, demand, unit_label)

    _sheet_demsat(wb, all_cases, q50, demand, frl_s, mddl_s, es_curve, unit_label)

    _sheet_cross_comparison(wb, all_cases, demand, unit_label)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
